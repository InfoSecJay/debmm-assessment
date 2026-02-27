#!/usr/bin/env python3
"""
DEBMM Assessment Spreadsheet Generator

Generates an all-in-one Excel spreadsheet with:
  - Tab 1: Instructions
  - Tab 2: Assessment (fillable with dropdowns and auto-scoring)
  - Tab 3: Results Dashboard (auto-calculated scores, heatmap, charts)
  - Tab 4: Rubric Reference

Usage:
    python generate_spreadsheet.py [--output debmm-assessment.xlsx]
    python generate_spreadsheet.py --mode audit --output audit-assessment.xlsx
"""

import argparse
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DEFAULT_RUBRIC = PROJECT_ROOT / "rubric" / "rubric.yaml"
DEFAULT_QUESTIONNAIRE = PROJECT_ROOT / "questionnaire" / "questionnaire.yaml"

# ── Color palette ─────────────────────────────────────────────────────────────

# DEBMM Core — navy/blue family
DARK_NAVY = "0F1D32"
NAVY = "1B2A4A"
STEEL = "2D3E50"
MED_BLUE = "3B82F6"
LIGHT_BLUE_BG = "EBF4FF"
BLUE_ACCENT = "93C5FD"

# Enrichment — teal family
TEAL = "0D7377"
DARK_TEAL = "115E60"
LIGHT_TEAL_BG = "E6F7F7"
TEAL_ACCENT = "5EEAD4"

# Neutrals
WHITE = "FFFFFF"
OFF_WHITE = "FAFBFC"
LIGHT_GRAY = "F1F5F9"
HAIRLINE_COLOR = "E2E8F0"
DARK_TEXT = "1E293B"
MED_TEXT = "475569"

# Answer cells — warm amber (the only warm tone)
ANSWER_BG_COLOR = "FFFBEB"
ANSWER_BORDER_COLOR = "F59E0B"

# Conditional formatting
SCORE_GREEN = "D1FAE5"
SCORE_YELLOW = "FEF3C7"
SCORE_ORANGE = "FED7AA"
SCORE_RED = "FEE2E2"

# Rubric level tints (5 distinct)
LEVEL_1_BG = "FFF1F2"
LEVEL_2_BG = "FFF7ED"
LEVEL_3_BG = "FEFCE8"
LEVEL_4_BG = "F0FDF4"
LEVEL_5_BG = "ECFDF5"

# ── Font constants ────────────────────────────────────────────────────────────

FN = "Aptos"  # Modern Excel 365 default; graceful fallback to Calibri

FONT_TITLE = Font(name=FN, size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name=FN, size=10, italic=True, color=BLUE_ACCENT)
FONT_SECTION = Font(name=FN, size=12, bold=True, color=NAVY)
FONT_SECTION_TEAL = Font(name=FN, size=12, bold=True, color=DARK_TEAL)
FONT_TIER_BANNER = Font(name=FN, size=12, bold=True, color=WHITE)
FONT_COL_HEADER = Font(name=FN, size=10, bold=True, color=WHITE)
FONT_BODY = Font(name=FN, size=10.5, color=DARK_TEXT)
FONT_BODY_BOLD = Font(name=FN, size=10.5, bold=True, color=DARK_TEXT)
FONT_BODY_ITALIC = Font(name=FN, size=10.5, italic=True, color=MED_TEXT)
FONT_SMALL = Font(name=FN, size=9.5, color=MED_TEXT)
FONT_SMALL_ITALIC = Font(name=FN, size=9.5, italic=True, color=MED_TEXT)
FONT_ANSWER = Font(name=FN, size=11, bold=True, color=DARK_TEXT)
FONT_SCORE_HERO = Font(name=FN, size=28, bold=True, color=NAVY)
FONT_SCORE_LARGE = Font(name=FN, size=16, bold=True, color=NAVY)
FONT_SCORE_LABEL = Font(name=FN, size=10, bold=True, color=STEEL)
FONT_LEVEL_BOLD = Font(name=FN, size=10, bold=True, color=DARK_TEXT)
FONT_CONTEXT = Font(name=FN, size=9.5, italic=True, color=MED_TEXT)

# ── Fill constants ────────────────────────────────────────────────────────────

FILL_DARK_NAVY = PatternFill("solid", fgColor=DARK_NAVY)
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_STEEL = PatternFill("solid", fgColor=STEEL)
FILL_TEAL = PatternFill("solid", fgColor=TEAL)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)
FILL_OFF_WHITE = PatternFill("solid", fgColor=OFF_WHITE)
FILL_LIGHT_GRAY = PatternFill("solid", fgColor=LIGHT_GRAY)
FILL_LIGHT_BLUE = PatternFill("solid", fgColor=LIGHT_BLUE_BG)
FILL_LIGHT_TEAL = PatternFill("solid", fgColor=LIGHT_TEAL_BG)
FILL_ANSWER = PatternFill("solid", fgColor=ANSWER_BG_COLOR)

FILL_SCORE_GREEN = PatternFill("solid", fgColor=SCORE_GREEN)
FILL_SCORE_YELLOW = PatternFill("solid", fgColor=SCORE_YELLOW)
FILL_SCORE_ORANGE = PatternFill("solid", fgColor=SCORE_ORANGE)
FILL_SCORE_RED = PatternFill("solid", fgColor=SCORE_RED)

FILL_LEVEL = {
    1: PatternFill("solid", fgColor=LEVEL_1_BG),
    2: PatternFill("solid", fgColor=LEVEL_2_BG),
    3: PatternFill("solid", fgColor=LEVEL_3_BG),
    4: PatternFill("solid", fgColor=LEVEL_4_BG),
    5: PatternFill("solid", fgColor=LEVEL_5_BG),
}

# ── Alignment constants ──────────────────────────────────────────────────────

ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ── Border constants ──────────────────────────────────────────────────────────

HAIRLINE_BOTTOM = Border(bottom=Side(style="thin", color=HAIRLINE_COLOR))
THIN_BORDER = Border(
    left=Side(style="thin", color=HAIRLINE_COLOR),
    right=Side(style="thin", color=HAIRLINE_COLOR),
    top=Side(style="thin", color=HAIRLINE_COLOR),
    bottom=Side(style="thin", color=HAIRLINE_COLOR),
)
ANSWER_BORDER = Border(
    left=Side(style="thin", color=ANSWER_BORDER_COLOR),
    right=Side(style="thin", color=ANSWER_BORDER_COLOR),
    top=Side(style="thin", color=ANSWER_BORDER_COLOR),
    bottom=Side(style="thin", color=ANSWER_BORDER_COLOR),
)
BLUE_ACCENT_LEFT = Border(left=Side(style="medium", color=MED_BLUE))
TEAL_ACCENT_LEFT = Border(left=Side(style="medium", color=TEAL_ACCENT))


# ── Helpers ───────────────────────────────────────────────────────────────────


def load_yaml(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None):
    for c in range(col_start, col_end + 1):
        style_cell(ws.cell(row=row, column=c), font, fill, alignment, border)


def is_enrichment(tier_id):
    return str(tier_id).startswith("enrichment")


def apply_conditional_formatting(ws, cell_range):
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["1", "1.49"], fill=FILL_SCORE_RED),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["1.5", "2.49"], fill=FILL_SCORE_ORANGE),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["2.5", "3.49"], fill=FILL_SCORE_YELLOW),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["3.5", "5"], fill=FILL_SCORE_GREEN),
    )


def level_formula(score_ref):
    return (
        f'=IF({score_ref}="","",IF({score_ref}>=4.5,"Optimized",'
        f'IF({score_ref}>=3.5,"Managed",IF({score_ref}>=2.5,"Defined",'
        f'IF({score_ref}>=1.5,"Repeatable","Initial")))))'
    )


def status_formula(score_ref):
    return f'=IF({score_ref}="","",IF({score_ref}>=3,"\u2713 Pass","\u2717 Below Target"))'


# ── Tab 1: Instructions ──────────────────────────────────────────────────────


def build_instructions_tab(wb: Workbook, mode: str):
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = NAVY

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 4

    # ── Title banner ──────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 56
    c = ws["A1"]
    c.value = "  DEBMM Assessment Tool"
    style_cell(c, FONT_TITLE, FILL_DARK_NAVY, Alignment(horizontal="left", vertical="center"))
    for col in "ABCD":
        ws[f"{col}1"].fill = FILL_DARK_NAVY

    # Subtitle row
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 26
    mode_label = "Self-Assessment" if mode == "self" else "Audit Assessment"
    c = ws["A2"]
    c.value = f"  Detection Engineering Behavior Maturity Model \u2014 {mode_label}"
    style_cell(c, FONT_SUBTITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    for col in "ABCD":
        ws[f"{col}2"].fill = FILL_NAVY

    row = 4

    def section_card(title, lines, accent="blue"):
        nonlocal row
        fill = FILL_LIGHT_BLUE if accent == "blue" else FILL_LIGHT_TEAL
        border = BLUE_ACCENT_LEFT if accent == "blue" else TEAL_ACCENT_LEFT
        font = FONT_SECTION if accent == "blue" else FONT_SECTION_TEAL

        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row=row, column=2, value=title)
        style_cell(ws.cell(row=row, column=2), font, fill, ALIGN_LEFT, border)
        ws.cell(row=row, column=3).fill = fill
        ws.row_dimensions[row].height = 28
        row += 1

        for line in lines:
            ws.merge_cells(f"B{row}:C{row}")
            ws.cell(row=row, column=2, value=line)
            style_cell(ws.cell(row=row, column=2), FONT_BODY, alignment=ALIGN_WRAP)
            row += 1

        row += 1  # Spacing

    # ── About ─────────────────────────────────────────────────────────
    section_card("About This Assessment", [
        "This spreadsheet assesses your detection engineering team's maturity using Elastic's "
        "Detection Engineering Behavior Maturity Model (DEBMM), enriched with organizational "
        "dimensions from detectionengineering.io.",
        "",
        "It covers 21 criteria across 7 categories, with 41 dropdown questions total.",
        "All questions use dropdowns (Yes/No or Scale 1\u20145). No free-text required.",
    ])

    # ── How to Use ────────────────────────────────────────────────────
    section_card("How to Use", [
        "1.  Go to the Assessment tab and fill in your organization details",
        "2.  Answer each question using the dropdown menus in the amber column",
        "3.  For Yes/No questions \u2014 select from the dropdown",
        "4.  For Scale questions (1\u20145) \u2014 select your maturity rating",
        "5.  Switch to the Results Dashboard tab to see scores calculated automatically",
    ])

    # ── Understanding the Model (two-column) ──────────────────────────
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="Understanding the Model")
    style_cell(ws.cell(row=row, column=2), FONT_SECTION, FILL_LIGHT_BLUE, ALIGN_LEFT, BLUE_ACCENT_LEFT)
    ws.cell(row=row, column=3).fill = FILL_LIGHT_BLUE
    ws.row_dimensions[row].height = 28
    row += 1

    # Left column header — DEBMM
    ws.cell(row=row, column=2, value="DEBMM Core Tiers (Elastic)")
    style_cell(ws.cell(row=row, column=2), FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_LEFT, BLUE_ACCENT_LEFT)
    # Right column header — Enrichment
    ws.cell(row=row, column=3, value="Supplementary Dimensions")
    style_cell(ws.cell(row=row, column=3), Font(name=FN, size=10.5, bold=True, color=DARK_TEAL),
               FILL_LIGHT_TEAL, ALIGN_LEFT, TEAL_ACCENT_LEFT)
    ws.row_dimensions[row].height = 24
    row += 1

    core_lines = [
        "Tier 0 \u2014 Foundation",
        "    Rule development, maintenance, roadmaps, threat modeling",
        "Tier 1 \u2014 Basic",
        "    Baseline rules, telemetry, version control, testing",
        "Tier 2 \u2014 Intermediate",
        "    FP reduction, gap analysis, internal validation",
        "Tier 3 \u2014 Advanced",
        "    FN triage, external validation, advanced TTP coverage",
        "Tier 4 \u2014 Expert",
        "    Threat hunting, automation, AI/LLM integration",
    ]
    enrichment_lines = [
        "People & Organization",
        "    Team structure, training, leadership",
        "Process & Governance",
        "    Lifecycle, metrics, collaboration",
        "",
        "These supplementary dimensions assess",
        "organizational readiness. They contribute",
        "to the overall score but do not affect",
        "DEBMM tier determination.",
        "",
    ]

    for i, (left, right) in enumerate(zip(core_lines, enrichment_lines)):
        ws.cell(row=row, column=2, value=left)
        style_cell(ws.cell(row=row, column=2), FONT_BODY, alignment=ALIGN_WRAP)
        ws.cell(row=row, column=3, value=right)
        style_cell(ws.cell(row=row, column=3), FONT_BODY, alignment=ALIGN_WRAP)
        row += 1

    row += 1  # Spacing

    # ── Maturity Levels ───────────────────────────────────────────────
    section_card("Maturity Levels", [
        "Each criterion is scored on a 1\u20145 scale:",
        "",
        "  1 \u2014 Initial          Minimal or no structured activity",
        "  2 \u2014 Repeatable    Sporadic, inconsistent efforts",
        "  3 \u2014 Defined         Regular, documented processes followed consistently",
        "  4 \u2014 Managed       Comprehensive, well-integrated with measurable outcomes",
        "  5 \u2014 Optimized     Fully automated, continuously improving",
    ])

    # ── Tier Determination ────────────────────────────────────────────
    section_card("Tier Determination", [
        "Your achieved tier is the highest tier where ALL criteria in that tier (and all lower "
        "tiers) score \u2265 3.0 (Defined level). This enforces the progressive nature of the "
        "model \u2014 you need solid foundations before claiming advanced maturity.",
        "",
        "Enrichment dimensions (People & Organization, Process & Governance) contribute to the "
        "overall maturity score but are not part of the tier determination logic.",
    ])

    # ── Scoring Paths ─────────────────────────────────────────────────
    section_card("Scoring Paths", [
        "1.  This Spreadsheet \u2014 Fill it out and scores calculate automatically in the Dashboard",
        "2.  Python CLI \u2014 Export answers to YAML and run scorer/score.py for a detailed report",
        "3.  LLM-Assisted \u2014 Run scorer/llm_scorer.py for AI-generated improvement recommendations",
    ])

    # ── References ────────────────────────────────────────────────────
    section_card("References", [
        "Elastic DEBMM: https://www.elastic.co/security-labs/elastic-releases-debmm",
        "Detection Engineering Maturity Matrix: https://detectionengineering.io/",
        "MITRE ATT&CK: https://attack.mitre.org/",
    ])

    return ws


# ── Tab 2: Assessment ─────────────────────────────────────────────────────────


def build_assessment_tab(wb: Workbook, questionnaire: dict, rubric: dict, mode: str):
    ws = wb.create_sheet("Assessment")
    ws.sheet_properties.tabColor = MED_BLUE

    # Column layout — gutter | ID | Criterion | Question | Answer | Score | [Evidence] | gutter
    if mode == "audit":
        col_widths = {"A": 2, "B": 7, "C": 26, "D": 78, "E": 14, "F": 10, "G": 45, "H": 2}
        last_col = "H"
        last_col_num = 8
        evidence_col = 7  # Column G
    else:
        col_widths = {"A": 2, "B": 7, "C": 26, "D": 78, "E": 14, "F": 10, "G": 2}
        last_col = "G"
        last_col_num = 7
        evidence_col = None
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Title banner ──────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 52
    c = ws["A1"]
    c.value = "  DEBMM Assessment"
    style_cell(c, FONT_TITLE, FILL_DARK_NAVY, Alignment(horizontal="left", vertical="center"))
    for i in range(1, last_col_num + 1):
        ws.cell(row=1, column=i).fill = FILL_DARK_NAVY

    # Subtitle row
    ws.merge_cells(f"A2:{last_col}2")
    ws.row_dimensions[2].height = 24
    c = ws["A2"]
    c.value = "  Detection Engineering Behavior Maturity Model"
    style_cell(c, FONT_SUBTITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    for i in range(1, last_col_num + 1):
        ws.cell(row=2, column=i).fill = FILL_NAVY

    # ── Metadata section ──────────────────────────────────────────────
    ws.row_dimensions[3].height = 8  # Spacer
    labels = ["Organization:", "Assessor Name:", "Assessor Role:", "Date:", "Assessment Type:"]
    for i, label in enumerate(labels):
        r = 4 + i
        ws.cell(row=r, column=3, value=label)
        style_cell(ws.cell(row=r, column=3), FONT_BODY_BOLD, alignment=ALIGN_RIGHT)
        ws.merge_cells(f"D{r}:E{r}")
        style_cell(ws.cell(row=r, column=4), FONT_BODY, FILL_ANSWER, ALIGN_LEFT, ANSWER_BORDER)
        ws.cell(row=r, column=5).fill = FILL_ANSWER
        ws.cell(row=r, column=5).border = ANSWER_BORDER

    # Pre-fill assessment type
    ws.cell(row=8, column=4, value="Self-Assessment" if mode == "self" else "Audit")

    # Type dropdown
    dv_type = DataValidation(type="list", formula1='"Self-Assessment,Audit"', allow_blank=False)
    dv_type.error = "Please select Self-Assessment or Audit"
    dv_type.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv_type)
    dv_type.add(ws.cell(row=8, column=4))

    # ── Data validations ──────────────────────────────────────────────
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yesno.error = "Please select Yes or No"
    dv_yesno.errorTitle = "Invalid Entry"
    dv_yesno.prompt = "Select Yes or No"
    dv_yesno.promptTitle = "Answer"
    ws.add_data_validation(dv_yesno)

    dv_scale = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_scale.error = "Please select a value from 1 to 5"
    dv_scale.errorTitle = "Invalid Entry"
    dv_scale.prompt = "Select 1 (Initial) through 5 (Optimized)"
    dv_scale.promptTitle = "Maturity Rating"
    ws.add_data_validation(dv_scale)

    # ── Build tier / criterion lookups ────────────────────────────────
    tier_names = {}
    tier_labels = {}
    tier_descs = {}
    for tier in rubric["tiers"]:
        tid = tier["id"]
        tier_names[tid] = tier["name"]
        desc = tier.get("description", "").strip()
        # Truncate long descriptions to first sentence
        if desc and ". " in desc:
            desc = desc[: desc.index(". ") + 1]
        tier_descs[tid] = desc
        if tid.startswith("tier_"):
            num = tid.replace("tier_", "")
            tier_names[int(num)] = tier["name"]
            tier_labels[int(num)] = f"TIER {num}: {tier['name'].upper()}"
            tier_labels[tid] = tier_labels[int(num)]
        else:
            tier_labels[tid] = tier["name"].upper().replace("ENRICHMENT: ", "")

    crit_names = {}
    for tier in rubric["tiers"]:
        for crit in tier["criteria"]:
            crit_names[crit["id"]] = crit["name"]

    # ── DEBMM Core Assessment context header ──────────────────────────
    row = 10
    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2, value="DEBMM CORE ASSESSMENT \u2014 Tiers 0\u20144")
    style_cell(ws.cell(row=row, column=2), FONT_SECTION, FILL_LIGHT_BLUE, ALIGN_LEFT, BLUE_ACCENT_LEFT)
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_BLUE
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2,
            value="Rate your team across the 5 progressive DEBMM tiers. "
                  "Your achieved tier is the highest where all criteria score \u2265 3.0.")
    style_cell(ws.cell(row=row, column=2), FONT_CONTEXT, FILL_LIGHT_BLUE, ALIGN_LEFT)
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_BLUE
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Column headers ────────────────────────────────────────────────
    if mode == "audit":
        headers = ["ID", "Criterion", "Question", "Answer", "Score", "Evidence / Notes"]
        hdr_cols = list(range(2, 8))
    else:
        headers = ["ID", "Criterion", "Question", "Answer", "Score"]
        hdr_cols = list(range(2, 7))

    header_row = row
    for col_idx, header in zip(hdr_cols, headers):
        c = ws.cell(row=header_row, column=col_idx, value=header)
        style_cell(c, FONT_COL_HEADER, FILL_STEEL, ALIGN_CENTER, THIN_BORDER)
    ws.row_dimensions[header_row].height = 26
    row += 1

    # ── Question rows ─────────────────────────────────────────────────
    current_tier = None
    question_rows = []
    q_key = "question" if mode == "self" else "question_audit"
    enrichment_headers_inserted = False
    prev_criterion = None
    q_idx_in_tier = 0  # For alternating rows

    for q in questionnaire["questions"]:
        q_tier = q["tier"]

        # ── Enrichment section transition ─────────────────────────────
        if is_enrichment(q_tier) and not enrichment_headers_inserted:
            enrichment_headers_inserted = True
            prev_criterion = None
            q_idx_in_tier = 0

            # Spacer / chapter break
            ws.row_dimensions[row].height = 20
            for c in range(1, last_col_num + 1):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_GRAY
            row += 1

            # Enrichment context header
            ws.merge_cells(f"B{row}:F{row}")
            ws.cell(row=row, column=2,
                    value="SUPPLEMENTARY DIMENSIONS \u2014 Organizational Readiness")
            style_cell(ws.cell(row=row, column=2), FONT_SECTION_TEAL, FILL_LIGHT_TEAL,
                       ALIGN_LEFT, TEAL_ACCENT_LEFT)
            for c in range(3, 7):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_TEAL
            ws.row_dimensions[row].height = 32
            row += 1

            ws.merge_cells(f"B{row}:F{row}")
            ws.cell(row=row, column=2,
                    value="These dimensions from detectionengineering.io assess people and process factors. "
                          "They contribute to the overall score but do not affect DEBMM tier determination.")
            style_cell(ws.cell(row=row, column=2), FONT_CONTEXT, FILL_LIGHT_TEAL, ALIGN_LEFT)
            for c in range(3, 7):
                ws.cell(row=row, column=c).fill = FILL_LIGHT_TEAL
            ws.row_dimensions[row].height = 22
            row += 1

            # Repeated column headers with teal
            for col_idx, hdr in zip(hdr_cols, headers):
                c = ws.cell(row=row, column=col_idx, value=hdr)
                style_cell(c, FONT_COL_HEADER, FILL_TEAL, ALIGN_CENTER, THIN_BORDER)
            ws.row_dimensions[row].height = 26
            row += 1

        # ── Tier separator banner ─────────────────────────────────────
        if q_tier != current_tier:
            current_tier = q_tier
            prev_criterion = None
            q_idx_in_tier = 0

            banner_fill = FILL_TEAL if is_enrichment(q_tier) else FILL_NAVY
            banner_text = tier_labels.get(q_tier, str(q_tier).upper())

            ws.merge_cells(f"B{row}:F{row}")
            ws.cell(row=row, column=2, value=f"  {banner_text}")
            style_cell(ws.cell(row=row, column=2), FONT_TIER_BANNER, banner_fill,
                       Alignment(horizontal="left", vertical="center"))
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = banner_fill
            if evidence_col:
                ws.cell(row=row, column=evidence_col).fill = banner_fill
            ws.row_dimensions[row].height = 34
            row += 1

            # Tier description sub-row
            desc = tier_descs.get(q_tier if isinstance(q_tier, str) else f"tier_{q_tier}", "")
            if not desc:
                desc = tier_descs.get(q_tier, "")
            if desc:
                sub_fill = FILL_LIGHT_TEAL if is_enrichment(q_tier) else FILL_LIGHT_BLUE
                ws.merge_cells(f"B{row}:F{row}")
                ws.cell(row=row, column=2, value=desc)
                style_cell(ws.cell(row=row, column=2), FONT_SMALL_ITALIC, sub_fill, ALIGN_LEFT)
                for c in range(3, 7):
                    ws.cell(row=row, column=c).fill = sub_fill
                ws.row_dimensions[row].height = 20
                row += 1

        # ── Question data ─────────────────────────────────────────────
        qid = q["id"]
        qtype = q["type"]
        criterion_id = q["criterion"]
        criterion = crit_names.get(criterion_id, criterion_id)
        question_text = q.get(q_key, q["question"])
        yes_value = q.get("scoring", {}).get("yes_value", 3) if qtype == "checklist" else None

        # Alternating row tint
        enrichment_q = is_enrichment(q_tier)
        if q_idx_in_tier % 2 == 1:
            row_fill = FILL_LIGHT_TEAL if enrichment_q else FILL_OFF_WHITE
        else:
            row_fill = FILL_WHITE

        # Criterion grouping — bold for first question of each criterion
        is_first_of_criterion = (criterion_id != prev_criterion)
        prev_criterion = criterion_id
        crit_font = FONT_BODY_BOLD if is_first_of_criterion else FONT_SMALL

        # Column B — ID
        ws.cell(row=row, column=2, value=qid)
        style_cell(ws.cell(row=row, column=2), FONT_SMALL, row_fill, ALIGN_CENTER, HAIRLINE_BOTTOM)

        # Column C — Criterion
        ws.cell(row=row, column=3, value=criterion)
        style_cell(ws.cell(row=row, column=3), crit_font, row_fill, ALIGN_LEFT_TOP, HAIRLINE_BOTTOM)

        # Column D — Question (with scale options)
        if qtype == "scale" and "options" in q:
            option_lines = "\n".join(f"{k} \u2014 {v}" for k, v in q["options"].items())
            full_question = f"{question_text}\n\n{option_lines}"
        else:
            full_question = question_text
        ws.cell(row=row, column=4, value=full_question)
        style_cell(ws.cell(row=row, column=4), FONT_BODY, row_fill, ALIGN_LEFT_TOP, HAIRLINE_BOTTOM)

        # Column E — Answer (warm amber)
        answer_cell = ws.cell(row=row, column=5)
        style_cell(answer_cell, FONT_ANSWER, FILL_ANSWER, ALIGN_CENTER, ANSWER_BORDER)
        if qtype == "checklist":
            dv_yesno.add(answer_cell)
        elif qtype == "scale":
            dv_scale.add(answer_cell)

        # Column F — Score (auto-calculated)
        score_cell = ws.cell(row=row, column=6)
        answer_ref = f"E{row}"
        if qtype == "checklist":
            score_cell.value = f'=IF({answer_ref}="Yes",{yes_value},IF({answer_ref}="No",1,""))'
        elif qtype == "scale":
            score_cell.value = f'=IF({answer_ref}="","",{answer_ref})'
        style_cell(score_cell, FONT_BODY_BOLD, row_fill, ALIGN_CENTER, HAIRLINE_BOTTOM)
        score_cell.number_format = "0.0"

        # Column G — Evidence (audit only)
        if evidence_col:
            evidence_cell = ws.cell(row=row, column=evidence_col)
            style_cell(evidence_cell, FONT_BODY, FILL_ANSWER, ALIGN_LEFT_TOP, ANSWER_BORDER)

        # Row height
        ws.row_dimensions[row].height = 115 if qtype == "scale" else 35

        question_rows.append({
            "row": row,
            "id": qid,
            "type": qtype,
            "criterion": criterion_id,
            "tier": q_tier,
            "yes_value": yes_value,
        })
        q_idx_in_tier += 1
        row += 1

    # Conditional formatting on score column (F)
    apply_conditional_formatting(ws, f"F{header_row + 1}:F{row - 1}")

    # Freeze panes below first column header row
    ws.freeze_panes = f"A{header_row + 1}"

    return ws, question_rows, header_row


# ── Tab 3: Results Dashboard ──────────────────────────────────────────────────


def build_dashboard_tab(wb: Workbook, rubric: dict, questionnaire: dict,
                        question_rows: list, header_row: int):
    ws = wb.create_sheet("Results Dashboard")
    ws.sheet_properties.tabColor = NAVY

    col_widths = {"A": 3, "B": 6, "C": 34, "D": 14, "E": 14, "F": 14, "G": 3}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    # Hidden chart data columns
    for c in ["M", "N", "O", "P"]:
        ws.column_dimensions[c].width = 18 if c in ("M", "O") else 10

    # ── Title banner ──────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 56
    c = ws["A1"]
    c.value = "  DEBMM Assessment Results"
    style_cell(c, FONT_TITLE, FILL_DARK_NAVY, Alignment(horizontal="left", vertical="center"))
    for col in "ABCDEFG":
        ws[f"{col}1"].fill = FILL_DARK_NAVY

    # Subtitle — org name
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 24
    c = ws["A2"]
    c.value = f'  =Assessment!D4'
    style_cell(c, FONT_SUBTITLE, FILL_NAVY, Alignment(horizontal="left", vertical="center"))
    for col in "ABCDEFG":
        ws[f"{col}2"].fill = FILL_NAVY

    # ── Executive Summary Cards ───────────────────────────────────────
    row = 5

    # Labels row
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="Overall Maturity Score")
    style_cell(ws.cell(row=row, column=2), FONT_SCORE_LABEL, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    ws.cell(row=row, column=3).fill = FILL_LIGHT_BLUE
    ws.cell(row=row, column=3).border = THIN_BORDER

    ws.merge_cells(f"D{row}:E{row}")
    ws.cell(row=row, column=4, value="Achieved DEBMM Tier")
    style_cell(ws.cell(row=row, column=4), FONT_SCORE_LABEL, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    ws.cell(row=row, column=5).fill = FILL_LIGHT_BLUE
    ws.cell(row=row, column=5).border = THIN_BORDER

    ws.cell(row=row, column=6, value="Completion")
    style_cell(ws.cell(row=row, column=6), FONT_SCORE_LABEL, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
    ws.row_dimensions[row].height = 22
    row += 1

    # Values row
    ws.row_dimensions[row].height = 60
    ws.merge_cells(f"B{row}:C{row}")
    style_cell(ws.cell(row=row, column=2), FONT_SCORE_HERO, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    ws.cell(row=row, column=3).border = THIN_BORDER

    ws.merge_cells(f"D{row}:E{row}")
    style_cell(ws.cell(row=row, column=4), FONT_SCORE_LARGE, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)
    ws.cell(row=row, column=5).border = THIN_BORDER

    total_q = len(question_rows)
    completion_formula = (
        f'=COUNTA(Assessment!E{question_rows[0]["row"]}:E{question_rows[-1]["row"]})'
        f'&" / {total_q}"'
    )
    ws.cell(row=row, column=6, value=completion_formula)
    style_cell(ws.cell(row=row, column=6), FONT_SCORE_LARGE, FILL_WHITE, ALIGN_CENTER, THIN_BORDER)

    overall_row = row
    row += 1

    # Explanation row
    ws.row_dimensions[row].height = 8  # Small spacer
    row += 1
    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2,
            value="Your achieved tier is the highest tier where all criteria in that tier "
                  "and all lower tiers score \u2265 3.0 (Defined level).")
    style_cell(ws.cell(row=row, column=2), FONT_SMALL_ITALIC, FILL_LIGHT_GRAY, ALIGN_LEFT)
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_GRAY
    ws.row_dimensions[row].height = 22
    row += 2

    # ── Build mappings ────────────────────────────────────────────────
    crit_to_rows = {}
    for qr in question_rows:
        crit_to_rows.setdefault(qr["criterion"], []).append(qr)

    tier_criteria_list = []
    for tier in rubric["tiers"]:
        for crit in tier["criteria"]:
            tier_criteria_list.append({
                "tier_id": tier["id"],
                "tier_name": tier["name"],
                "crit_id": crit["id"],
                "crit_name": crit["name"],
            })

    core_tier_ids = {"tier_0", "tier_1", "tier_2", "tier_3", "tier_4"}
    core_criteria = [tc for tc in tier_criteria_list if tc["tier_id"] in core_tier_ids]
    enrichment_criteria = [tc for tc in tier_criteria_list if tc["tier_id"] not in core_tier_ids]

    # ── DEBMM Core Section ────────────────────────────────────────────
    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2, value="  DEBMM CORE ASSESSMENT")
    style_cell(ws.cell(row=row, column=2), FONT_TIER_BANNER, FILL_NAVY,
               Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = FILL_NAVY
    ws.row_dimensions[row].height = 30
    row += 1

    # Column headers
    core_hdr_row = row
    for col_idx, hdr in [(3, "Category / Criterion"), (4, "Score"), (5, "Level"), (6, "Status")]:
        ws.cell(row=row, column=col_idx, value=hdr)
        style_cell(ws.cell(row=row, column=col_idx), FONT_COL_HEADER, FILL_STEEL, ALIGN_CENTER, THIN_BORDER)
    ws.row_dimensions[row].height = 24
    row += 1

    criterion_score_cells = []
    core_criterion_cells = []
    tier_score_cells = {}
    current_tier_id = None
    tier_crit_cells = []

    def _finalize_tier(tid):
        nonlocal tier_crit_cells
        if tid and tier_crit_cells and tid in tier_score_cells:
            trow = tier_score_cells[tid]["row"]
            avg_refs = ",".join(tier_crit_cells)
            ws.cell(row=trow, column=4, value=f'=IF(COUNT({avg_refs})=0,"",AVERAGE({avg_refs}))')
            ws.cell(row=trow, column=4).number_format = "0.00"
            ws.cell(row=trow, column=5, value=level_formula(f"D{trow}"))
            ws.cell(row=trow, column=6, value=status_formula(f"D{trow}"))
        tier_crit_cells = []

    for tc in core_criteria:
        if tc["tier_id"] != current_tier_id:
            _finalize_tier(current_tier_id)
            current_tier_id = tc["tier_id"]

            # Tier summary row
            ws.cell(row=row, column=3, value=tc["tier_name"])
            style_cell(ws.cell(row=row, column=3), FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_LEFT, THIN_BORDER)
            for c in range(4, 7):
                style_cell(ws.cell(row=row, column=c), FONT_BODY_BOLD, FILL_LIGHT_BLUE, ALIGN_CENTER, THIN_BORDER)
            tier_score_cells[current_tier_id] = {"row": row, "name": tc["tier_name"]}
            ws.row_dimensions[row].height = 28
            row += 1

        # Criterion row
        crit_id = tc["crit_id"]
        rows_for_crit = crit_to_rows.get(crit_id, [])

        ws.cell(row=row, column=3, value=f"    {tc['crit_name']}")
        style_cell(ws.cell(row=row, column=3), FONT_BODY, FILL_WHITE, ALIGN_LEFT, HAIRLINE_BOTTOM)

        if rows_for_crit:
            score_refs = [f"Assessment!F{qr['row']}" for qr in rows_for_crit]
            refs_str = ",".join(score_refs)
            ws.cell(row=row, column=4, value=f'=IF(COUNT({refs_str})=0,"",AVERAGE({refs_str}))')
        ws.cell(row=row, column=4).number_format = "0.00"
        style_cell(ws.cell(row=row, column=4), FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, HAIRLINE_BOTTOM)

        ws.cell(row=row, column=5, value=level_formula(f"D{row}"))
        style_cell(ws.cell(row=row, column=5), FONT_BODY, FILL_WHITE, ALIGN_CENTER, HAIRLINE_BOTTOM)

        ws.cell(row=row, column=6, value=status_formula(f"D{row}"))
        style_cell(ws.cell(row=row, column=6), FONT_BODY, FILL_WHITE, ALIGN_CENTER, HAIRLINE_BOTTOM)

        criterion_score_cells.append(f"D{row}")
        core_criterion_cells.append(f"D{row}")
        tier_crit_cells.append(f"D{row}")
        row += 1

    _finalize_tier(current_tier_id)

    # Conditional formatting on core section
    core_end_row = row - 1
    apply_conditional_formatting(ws, f"D{core_hdr_row + 1}:D{core_end_row}")
    ws.conditional_formatting.add(
        f"F{core_hdr_row + 1}:F{core_end_row}",
        CellIsRule(operator="equal", formula=['"✓ Pass"'], fill=FILL_SCORE_GREEN),
    )
    ws.conditional_formatting.add(
        f"F{core_hdr_row + 1}:F{core_end_row}",
        CellIsRule(operator="equal", formula=['"✗ Below Target"'], fill=FILL_SCORE_RED),
    )

    # ── Spacer ────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_GRAY
    row += 1

    # ── Enrichment Section ────────────────────────────────────────────
    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2, value="  SUPPLEMENTARY DIMENSIONS")
    style_cell(ws.cell(row=row, column=2), FONT_TIER_BANNER, FILL_TEAL,
               Alignment(horizontal="left", vertical="center"))
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = FILL_TEAL
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2,
            value="Organizational readiness factors \u2014 do not affect DEBMM tier determination")
    style_cell(ws.cell(row=row, column=2), FONT_SMALL_ITALIC, FILL_LIGHT_TEAL, ALIGN_LEFT)
    for c in range(3, 7):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_TEAL
    ws.row_dimensions[row].height = 20
    row += 1

    # Column headers (teal)
    enrich_hdr_row = row
    for col_idx, hdr in [(3, "Category / Criterion"), (4, "Score"), (5, "Level"), (6, "Status")]:
        ws.cell(row=row, column=col_idx, value=hdr)
        style_cell(ws.cell(row=row, column=col_idx), FONT_COL_HEADER, FILL_TEAL, ALIGN_CENTER, THIN_BORDER)
    ws.row_dimensions[row].height = 24
    row += 1

    enrich_tier_score_cells = {}
    current_tier_id = None
    tier_crit_cells = []

    for tc in enrichment_criteria:
        if tc["tier_id"] != current_tier_id:
            _finalize_tier(current_tier_id)
            current_tier_id = tc["tier_id"]

            # Category summary row
            display_name = tc["tier_name"].replace("Enrichment: ", "")
            ws.cell(row=row, column=3, value=display_name)
            style_cell(ws.cell(row=row, column=3), FONT_BODY_BOLD, FILL_LIGHT_TEAL, ALIGN_LEFT, THIN_BORDER)
            for c in range(4, 7):
                style_cell(ws.cell(row=row, column=c), FONT_BODY_BOLD, FILL_LIGHT_TEAL, ALIGN_CENTER, THIN_BORDER)
            tier_score_cells[current_tier_id] = {"row": row, "name": display_name}
            enrich_tier_score_cells[current_tier_id] = {"row": row, "name": display_name}
            ws.row_dimensions[row].height = 28
            row += 1

        # Criterion row
        crit_id = tc["crit_id"]
        rows_for_crit = crit_to_rows.get(crit_id, [])

        ws.cell(row=row, column=3, value=f"    {tc['crit_name']}")
        style_cell(ws.cell(row=row, column=3), FONT_BODY, FILL_WHITE, ALIGN_LEFT, HAIRLINE_BOTTOM)

        if rows_for_crit:
            score_refs = [f"Assessment!F{qr['row']}" for qr in rows_for_crit]
            refs_str = ",".join(score_refs)
            ws.cell(row=row, column=4, value=f'=IF(COUNT({refs_str})=0,"",AVERAGE({refs_str}))')
        ws.cell(row=row, column=4).number_format = "0.00"
        style_cell(ws.cell(row=row, column=4), FONT_BODY_BOLD, FILL_WHITE, ALIGN_CENTER, HAIRLINE_BOTTOM)

        ws.cell(row=row, column=5, value=level_formula(f"D{row}"))
        style_cell(ws.cell(row=row, column=5), FONT_BODY, FILL_WHITE, ALIGN_CENTER, HAIRLINE_BOTTOM)

        ws.cell(row=row, column=6, value=status_formula(f"D{row}"))
        style_cell(ws.cell(row=row, column=6), FONT_BODY, FILL_WHITE, ALIGN_CENTER, HAIRLINE_BOTTOM)

        criterion_score_cells.append(f"D{row}")
        tier_crit_cells.append(f"D{row}")
        row += 1

    _finalize_tier(current_tier_id)

    # Conditional formatting on enrichment section
    enrich_end_row = row - 1
    apply_conditional_formatting(ws, f"D{enrich_hdr_row + 1}:D{enrich_end_row}")
    ws.conditional_formatting.add(
        f"F{enrich_hdr_row + 1}:F{enrich_end_row}",
        CellIsRule(operator="equal", formula=['"✓ Pass"'], fill=FILL_SCORE_GREEN),
    )
    ws.conditional_formatting.add(
        f"F{enrich_hdr_row + 1}:F{enrich_end_row}",
        CellIsRule(operator="equal", formula=['"✗ Below Target"'], fill=FILL_SCORE_RED),
    )

    # ── Overall score formula ─────────────────────────────────────────
    if criterion_score_cells:
        all_refs = ",".join(criterion_score_cells)
        ws.cell(row=overall_row, column=2,
                value=f'=IF(COUNT({all_refs})=0,"",ROUND(AVERAGE({all_refs}),2)&" / 5.0")')
        ws.cell(row=overall_row, column=2).number_format = "@"

    # ── Achieved tier formula ─────────────────────────────────────────
    core_tiers_ordered = ["tier_0", "tier_1", "tier_2", "tier_3", "tier_4"]
    tier_display_labels = {
        "tier_0": "Tier 0: Foundation",
        "tier_1": "Tier 1: Basic",
        "tier_2": "Tier 2: Intermediate",
        "tier_3": "Tier 3: Advanced",
        "tier_4": "Tier 4: Expert",
    }

    tier_check_parts = {tid: [] for tid in core_tiers_ordered}
    crit_cell_idx = 0
    current_tier_scan = None
    for tc_item in core_criteria:
        if tc_item["tier_id"] != current_tier_scan:
            current_tier_scan = tc_item["tier_id"]
        if crit_cell_idx < len(core_criterion_cells):
            tier_check_parts[current_tier_scan].append(core_criterion_cells[crit_cell_idx])
        crit_cell_idx += 1

    def tier_check(tid):
        cells = tier_check_parts.get(tid, [])
        if not cells:
            return "TRUE"
        return f"AND({','.join(f'{c}>=3' for c in cells)})"

    cumul = {}
    for i, tid in enumerate(core_tiers_ordered):
        checks = [tier_check(core_tiers_ordered[j]) for j in range(i + 1)]
        cumul[tid] = f"AND({','.join(checks)})"

    tier_formula = f'=IF({cumul["tier_4"]},"{tier_display_labels["tier_4"]}",'
    tier_formula += f'IF({cumul["tier_3"]},"{tier_display_labels["tier_3"]}",'
    tier_formula += f'IF({cumul["tier_2"]},"{tier_display_labels["tier_2"]}",'
    tier_formula += f'IF({cumul["tier_1"]},"{tier_display_labels["tier_1"]}",'
    tier_formula += f'IF({cumul["tier_0"]},"{tier_display_labels["tier_0"]}",'
    tier_formula += '"Below Foundation")))))'
    ws.cell(row=overall_row, column=4, value=tier_formula)

    # Conditional formatting on overall score
    apply_conditional_formatting(ws, f"B{overall_row}:C{overall_row}")

    # ── Charts ────────────────────────────────────────────────────────
    row += 2

    # Chart 1: DEBMM Tier Scores (columns M-N)
    core_tier_row_list = [
        (ts["name"], ts["row"])
        for tid, ts in tier_score_cells.items()
        if tid in core_tier_ids
    ]
    if core_tier_row_list:
        chart_row = 2
        ws.cell(row=chart_row, column=13, value="Tier")
        ws.cell(row=chart_row, column=14, value="Score")
        chart_row += 1
        for name, trow in core_tier_row_list:
            ws.cell(row=chart_row, column=13, value=name)
            ws.cell(row=chart_row, column=14, value=f"=D{trow}")
            ws.cell(row=chart_row, column=14).number_format = "0.00"
            chart_row += 1

        # Target reference line data
        target_start = chart_row
        ws.cell(row=chart_row, column=13, value="Target Tier")
        ws.cell(row=chart_row, column=14, value="Target")
        chart_row += 1
        for name, _ in core_tier_row_list:
            ws.cell(row=chart_row, column=13, value=name)
            ws.cell(row=chart_row, column=14, value=3.0)
            chart_row += 1

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "DEBMM Tier Scores"
        chart1.y_axis.title = "Score (1\u20145)"
        chart1.y_axis.scaling.min = 0
        chart1.y_axis.scaling.max = 5
        chart1.x_axis.title = None
        chart1.legend = None

        data1 = Reference(ws, min_col=14, min_row=2, max_row=2 + len(core_tier_row_list))
        cats1 = Reference(ws, min_col=13, min_row=3, max_row=2 + len(core_tier_row_list))
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)

        # Color the bars MED_BLUE
        if chart1.series:
            for pt_idx in range(len(core_tier_row_list)):
                pt = DataPoint(idx=pt_idx)
                pt.graphicalProperties.solidFill = MED_BLUE
                chart1.series[0].data_points.append(pt)

        chart1.width = 18
        chart1.height = 13
        ws.add_chart(chart1, f"B{row}")

    # Chart 2: Enrichment Scores (columns O-P)
    enrich_row_list = [
        (ts["name"], ts["row"])
        for ts in enrich_tier_score_cells.values()
    ]
    if enrich_row_list:
        chart_row2 = 2
        ws.cell(row=chart_row2, column=15, value="Dimension")
        ws.cell(row=chart_row2, column=16, value="Score")
        chart_row2 += 1
        for name, trow in enrich_row_list:
            ws.cell(row=chart_row2, column=15, value=name)
            ws.cell(row=chart_row2, column=16, value=f"=D{trow}")
            ws.cell(row=chart_row2, column=16).number_format = "0.00"
            chart_row2 += 1

        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Organizational Readiness"
        chart2.y_axis.title = "Score (1\u20145)"
        chart2.y_axis.scaling.min = 0
        chart2.y_axis.scaling.max = 5
        chart2.x_axis.title = None
        chart2.legend = None

        data2 = Reference(ws, min_col=16, min_row=2, max_row=2 + len(enrich_row_list))
        cats2 = Reference(ws, min_col=15, min_row=3, max_row=2 + len(enrich_row_list))
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)

        if chart2.series:
            for pt_idx in range(len(enrich_row_list)):
                pt = DataPoint(idx=pt_idx)
                pt.graphicalProperties.solidFill = TEAL
                chart2.series[0].data_points.append(pt)

        chart2.width = 12
        chart2.height = 13
        ws.add_chart(chart2, f"E{row}")

    ws.freeze_panes = "A3"
    return ws


# ── Tab 4: Rubric Reference ──────────────────────────────────────────────────


def build_rubric_tab(wb: Workbook, rubric: dict):
    ws = wb.create_sheet("Rubric Reference")
    ws.sheet_properties.tabColor = STEEL

    col_widths = {"A": 14, "B": 60, "C": 40}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 48
    c = ws["A1"]
    c.value = "  DEBMM Rubric Reference"
    style_cell(c, FONT_TITLE, FILL_DARK_NAVY, Alignment(horizontal="left", vertical="center"))
    for col in "ABC":
        ws[f"{col}1"].fill = FILL_DARK_NAVY

    row = 3
    level_names = {1: "Initial", 2: "Repeatable", 3: "Defined", 4: "Managed", 5: "Optimized"}

    core_tiers = [t for t in rubric["tiers"] if not is_enrichment(t["id"])]
    enrichment_tiers = [t for t in rubric["tiers"] if is_enrichment(t["id"])]

    def render_tiers(tiers, section_label, banner_fill, section_fill, accent_border, section_font):
        nonlocal row

        # Section header
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1, value=f"  {section_label}")
        style_cell(ws.cell(row=row, column=1), FONT_TIER_BANNER, banner_fill,
                   Alignment(horizontal="left", vertical="center"))
        for c in range(2, 4):
            ws.cell(row=row, column=c).fill = banner_fill
        ws.row_dimensions[row].height = 28
        row += 1

        for tier in tiers:
            # Tier banner
            ws.merge_cells(f"A{row}:C{row}")
            tier_name = tier["name"]
            if tier["id"].startswith("tier_"):
                num = tier["id"].replace("tier_", "")
                tier_name = f"Tier {num}: {tier['name']}"
            else:
                tier_name = tier["name"].replace("Enrichment: ", "")
            ws.cell(row=row, column=1, value=f"  {tier_name}")
            style_cell(ws.cell(row=row, column=1), FONT_TIER_BANNER, banner_fill,
                       Alignment(horizontal="left", vertical="center"))
            for c in range(2, 4):
                ws.cell(row=row, column=c).fill = banner_fill
            ws.row_dimensions[row].height = 30
            row += 1

            # Tier description
            desc = tier.get("description", "").strip()
            if desc:
                ws.merge_cells(f"A{row}:C{row}")
                ws.cell(row=row, column=1, value=desc)
                style_cell(ws.cell(row=row, column=1), FONT_SMALL_ITALIC, section_fill, ALIGN_WRAP)
                for c in range(2, 4):
                    ws.cell(row=row, column=c).fill = section_fill
                ws.row_dimensions[row].height = 36
                row += 1

            for crit in tier["criteria"]:
                # Criterion name
                ws.merge_cells(f"A{row}:C{row}")
                ws.cell(row=row, column=1, value=crit["name"])
                style_cell(ws.cell(row=row, column=1), section_font, section_fill, ALIGN_LEFT, accent_border)
                for c in range(2, 4):
                    ws.cell(row=row, column=c).fill = section_fill
                ws.row_dimensions[row].height = 26
                row += 1

                # Sub-headers
                for col_idx, hdr in enumerate(["Level", "Description", "Quantitative Measure"], 1):
                    ws.cell(row=row, column=col_idx, value=hdr)
                    style_cell(ws.cell(row=row, column=col_idx), FONT_COL_HEADER, FILL_STEEL, ALIGN_CENTER, THIN_BORDER)
                ws.row_dimensions[row].height = 22
                row += 1

                for level_num in sorted(crit["levels"].keys()):
                    level_data = crit["levels"][level_num]
                    row_fill = FILL_LEVEL.get(level_num, FILL_WHITE)

                    ws.cell(row=row, column=1, value=f"{level_num} \u2014 {level_names[level_num]}")
                    style_cell(ws.cell(row=row, column=1), FONT_LEVEL_BOLD, row_fill, ALIGN_CENTER, THIN_BORDER)

                    ws.cell(row=row, column=2, value=level_data["qualitative"].strip())
                    style_cell(ws.cell(row=row, column=2), FONT_BODY, row_fill, ALIGN_WRAP, THIN_BORDER)

                    ws.cell(row=row, column=3, value=level_data.get("quantitative", "").strip())
                    style_cell(ws.cell(row=row, column=3), FONT_BODY, row_fill, ALIGN_WRAP, THIN_BORDER)

                    ws.row_dimensions[row].height = 42
                    row += 1

                row += 1  # Space between criteria

    # Render core tiers
    render_tiers(core_tiers, "DEBMM CORE TIERS", FILL_NAVY, FILL_LIGHT_BLUE,
                 BLUE_ACCENT_LEFT, FONT_SECTION)

    # Spacer
    ws.row_dimensions[row].height = 16
    row += 1

    # Render enrichment tiers
    render_tiers(enrichment_tiers, "SUPPLEMENTARY DIMENSIONS", FILL_TEAL, FILL_LIGHT_TEAL,
                 TEAL_ACCENT_LEFT, FONT_SECTION_TEAL)

    ws.freeze_panes = "A2"
    return ws


# ── Main ──────────────────────────────────────────────────────────────────────


def generate_spreadsheet(
    output_path: Path,
    rubric_path: Path = DEFAULT_RUBRIC,
    questionnaire_path: Path = DEFAULT_QUESTIONNAIRE,
    mode: str = "self",
):
    rubric = load_yaml(rubric_path)
    questionnaire = load_yaml(questionnaire_path)

    wb = Workbook()
    build_instructions_tab(wb, mode)
    _, question_rows, header_row = build_assessment_tab(wb, questionnaire, rubric, mode)
    build_dashboard_tab(wb, rubric, questionnaire, question_rows, header_row)
    build_rubric_tab(wb, rubric)

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate DEBMM assessment spreadsheet.")
    parser.add_argument(
        "--output", "-o", type=Path,
        default=PROJECT_ROOT / "templates" / "debmm-assessment.xlsx",
        help="Output Excel file path (default: templates/debmm-assessment.xlsx)",
    )
    parser.add_argument(
        "--mode", choices=["self", "audit"], default="self",
        help="Assessment mode: self-assessment or audit (default: self)",
    )
    parser.add_argument("--rubric", type=Path, default=DEFAULT_RUBRIC)
    parser.add_argument("--questionnaire", type=Path, default=DEFAULT_QUESTIONNAIRE)

    args = parser.parse_args()
    output = generate_spreadsheet(args.output, args.rubric, args.questionnaire, args.mode)
    print(f"Spreadsheet generated: {output}")


if __name__ == "__main__":
    main()
