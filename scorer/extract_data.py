"""Extract DEBMM assessment data from a completed .xlsx into JSON for report generation.

Usage:
    python scorer/extract_data.py <assessment.xlsx> [-o output.json] [--history history.json] [--date YYYY-MM]

The spreadsheet MUST be opened and saved in Excel first so that formulas are evaluated.
Reads the Report Data tab (designed for machine consumption).
"""

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


def _to_float(value):
    """Safely convert a cell value to float."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _to_str(value, default=""):
    """Safely convert a cell value to string."""
    if value is None:
        return default
    return str(value).strip()


def extract_report_data(xlsx_path: Path) -> dict:
    """Read the Report Data tab and return structured dict matching generate_report.js schema."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Report Data" not in wb.sheetnames:
        print(f"Error: No 'Report Data' tab found in {xlsx_path}", file=sys.stderr)
        print(f"Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb["Report Data"]

    # ── Table 1: Summary (rows 2-8, columns A-B) ──
    summary = {}
    for row in range(2, 9):
        label = _to_str(ws.cell(row=row, column=1).value)
        value = ws.cell(row=row, column=2).value
        summary[label] = value

    org = _to_str(summary.get("Organization"))
    assessor = _to_str(summary.get("Assessor"))
    date_val = summary.get("Date")
    if hasattr(date_val, "strftime"):
        date_str = date_val.strftime("%Y-%m-%d")
    else:
        date_str = _to_str(date_val, "N/A")
    assess_type = _to_str(summary.get("Assessment Type", "Self-Assessment"))

    overall_raw = summary.get("Overall Score")
    overall_score = _to_float(overall_raw)
    if overall_score is None and isinstance(overall_raw, str):
        try:
            overall_score = float(overall_raw.split("/")[0].strip())
        except (ValueError, IndexError):
            overall_score = 0.0

    achieved_tier = _to_str(summary.get("Achieved Tier", "N/A"))
    completion = _to_str(summary.get("Completion", "0 / 0"))

    # ── Table 2: Tier Progression (row 10 = header, rows 11-15 = data) ──
    tiers = []
    for row in range(11, 16):
        tier_id = _to_str(ws.cell(row=row, column=1).value)
        if not tier_id:
            break
        tier_name = _to_str(ws.cell(row=row, column=2).value)
        score = _to_float(ws.cell(row=row, column=3).value)
        level = _to_str(ws.cell(row=row, column=4).value, "N/A")
        status = _to_str(ws.cell(row=row, column=5).value, "N/A")
        progression = _to_str(ws.cell(row=row, column=6).value, "Not Started")

        tiers.append({
            "id": tier_id,
            "name": tier_name,
            "score": score if score is not None else 0.0,
            "level": level,
            "status": status,
            "progression": progression,
        })

    # ── Table 3: Criterion Breakdown (row 17 = header, rows 18+ = data) ──
    criteria = []
    row = 18
    while True:
        section = _to_str(ws.cell(row=row, column=1).value)
        if not section:
            break
        category = _to_str(ws.cell(row=row, column=2).value)
        criterion = _to_str(ws.cell(row=row, column=3).value)
        score = _to_float(ws.cell(row=row, column=4).value)
        level = _to_str(ws.cell(row=row, column=5).value, "N/A")
        status = _to_str(ws.cell(row=row, column=6).value, "N/A")

        criteria.append({
            "section": section,
            "category": category,
            "criterion": criterion,
            "score": score if score is not None else 0.0,
            "level": level,
            "status": status,
        })
        row += 1

    return {
        "org": org,
        "assessor": assessor,
        "date": date_str,
        "type": assess_type,
        "overallScore": overall_score if overall_score is not None else 0.0,
        "achievedTier": achieved_tier,
        "completion": completion,
        "tiers": tiers,
        "criteria": criteria,
    }


def _derive_period(data: dict, date_override: str | None) -> str:
    """Derive the YYYY-MM period key from --date override, spreadsheet date, or current month."""
    if date_override:
        return date_override

    date_str = data.get("date", "")
    if date_str and date_str != "N/A" and date_str != "0":
        try:
            # Handle YYYY-MM-DD
            dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
            return dt.strftime("%Y-%m")
        except ValueError:
            pass
        try:
            # Handle YYYY-MM directly
            datetime.strptime(date_str[:7], "%Y-%m")
            return date_str[:7]
        except ValueError:
            pass

    return datetime.now().strftime("%Y-%m")


def upsert_history(history_path: Path, data: dict, period: str):
    """Append or replace an entry in the history file, keyed by period (YYYY-MM)."""
    if history_path.exists():
        history = json.loads(history_path.read_text(encoding="utf-8"))
    else:
        history = []

    snapshot = {
        "date": period,
        "extractedAt": datetime.now(timezone.utc).isoformat(),
        **data,
    }

    # Upsert: replace existing entry for this period, or append
    replaced = False
    for i, entry in enumerate(history):
        if entry.get("date") == period:
            history[i] = snapshot
            replaced = True
            break
    if not replaced:
        history.append(snapshot)

    # Sort chronologically
    history.sort(key=lambda e: e.get("date", ""))

    history_path.write_text(
        json.dumps(history, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    action = "Updated" if replaced else "Added"
    print(f"  History: {action} period {period} in {history_path} ({len(history)} total entries)")


def main():
    parser = argparse.ArgumentParser(
        description="Extract DEBMM assessment data from xlsx to JSON."
    )
    parser.add_argument("xlsx", type=Path, help="Path to completed assessment .xlsx")
    parser.add_argument(
        "-o", "--output", type=Path, default=None,
        help="Output JSON path (default: <input>_data.json)",
    )
    parser.add_argument(
        "--history", type=Path, default=None,
        help="Path to history.json — appends/upserts this assessment for trend reporting",
    )
    parser.add_argument(
        "--date", type=str, default=None,
        help="Override assessment period (YYYY-MM format, e.g. 2026-03). "
             "Default: derived from spreadsheet date field or current month.",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"Error: File not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    if args.date:
        try:
            datetime.strptime(args.date, "%Y-%m")
        except ValueError:
            print(f"Error: --date must be YYYY-MM format (got: {args.date})", file=sys.stderr)
            sys.exit(1)

    data = extract_report_data(args.xlsx)

    # Write single-extract JSON
    output = args.output or args.xlsx.with_suffix("").with_name(
        args.xlsx.stem + "_data.json"
    )
    output.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Data extracted: {output}")
    print(f"  Organization: {data['org'] or 'N/A'}")
    print(f"  Overall Score: {data['overallScore']}")
    print(f"  Achieved Tier: {data['achievedTier']}")
    print(f"  Tiers: {len(data['tiers'])}, Criteria: {len(data['criteria'])}")

    # Append to history if requested
    if args.history:
        period = _derive_period(data, args.date)
        upsert_history(args.history, data, period)


if __name__ == "__main__":
    main()
