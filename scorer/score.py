#!/usr/bin/env python3
"""
DEBMM Assessment Scorer

Scores a completed DEBMM assessment response against the rubric.
Handles checklist (yes/no) and scale (1-5) questions automatically.
Flags text questions for manual or LLM review.

Usage:
    python score.py <response.yaml> [--rubric <rubric.yaml>] [--questionnaire <questionnaire.yaml>]
    python score.py <response.yaml> --json          # Output raw JSON
    python score.py <response.yaml> --report out.md # Generate markdown report
    python score.py --from-xlsx filled.xlsx          # Score from a filled-out spreadsheet
"""

import argparse
import json
import sys
from pathlib import Path

import yaml

# Default paths relative to this script
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
DEFAULT_RUBRIC = PROJECT_ROOT / "rubric" / "rubric.yaml"
DEFAULT_QUESTIONNAIRE = PROJECT_ROOT / "questionnaire" / "questionnaire.yaml"


def load_yaml(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_xlsx_responses(xlsx_path: Path, questionnaire_path: Path) -> dict:
    """Parse a filled-out DEBMM assessment spreadsheet into a response dict.

    Reads the Assessment tab and extracts answers by matching question IDs
    in column A to answer values in column F (Your Answer).
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Assessment"]

    questionnaire = load_yaml(questionnaire_path)
    q_index = {q["id"]: q for q in questionnaire["questions"]}

    # Read metadata from fixed cells
    metadata = {
        "organization": ws["C3"].value or "",
        "assessor_name": ws["C4"].value or "",
        "assessor_role": ws["C5"].value or "",
        "date": str(ws["C6"].value or ""),
        "assessment_type": str(ws["C7"].value or "self").lower().replace("-", "_").replace("self_assessment", "self"),
    }

    # Scan rows for question IDs in column A, answers in column F, evidence in column H
    responses = {}
    for row in ws.iter_rows(min_row=2, max_col=8):
        cell_a = row[0].value  # Column A: ID
        cell_f = row[5].value  # Column F: Answer
        cell_h = row[7].value  # Column H: Evidence

        if cell_a is None or cell_a not in q_index:
            continue

        qid = cell_a
        q_def = q_index[qid]
        answer = cell_f
        evidence = cell_h or ""

        # Convert answer based on question type
        if q_def["type"] == "checklist":
            if isinstance(answer, str):
                answer = answer.strip().lower() in ("yes", "true", "1")
            elif isinstance(answer, bool):
                pass
            else:
                answer = None
        elif q_def["type"] == "scale":
            if isinstance(answer, (int, float)):
                answer = int(answer)
            elif isinstance(answer, str) and answer.strip().isdigit():
                answer = int(answer.strip())
            else:
                answer = None
        elif q_def["type"] == "text":
            answer = str(answer).strip() if answer else ""

        responses[qid] = {
            "answer": answer,
            "evidence": evidence,
        }

    wb.close()

    return {"metadata": metadata, "responses": responses}


def build_question_index(questionnaire: dict) -> dict:
    """Build a lookup of question ID -> question definition."""
    return {q["id"]: q for q in questionnaire["questions"]}


def build_criterion_index(rubric: dict) -> dict:
    """Build a lookup of criterion ID -> criterion definition with tier info."""
    index = {}
    for tier in rubric["tiers"]:
        for criterion in tier["criteria"]:
            index[criterion["id"]] = {
                **criterion,
                "tier_id": tier["id"],
                "tier_name": tier["name"],
            }
    return index


def score_question(question: dict, answer) -> dict:
    """Score a single question response. Returns score info dict."""
    qtype = question["type"]
    qid = question["id"]

    if answer is None:
        return {
            "id": qid,
            "type": qtype,
            "score": None,
            "status": "unanswered",
            "raw_answer": None,
        }

    if qtype == "checklist":
        if isinstance(answer, bool):
            yes_value = question["scoring"]["yes_value"]
            score = yes_value if answer else 1
            return {
                "id": qid,
                "type": qtype,
                "score": score,
                "status": "scored",
                "raw_answer": answer,
            }
        return {
            "id": qid,
            "type": qtype,
            "score": None,
            "status": "invalid",
            "raw_answer": answer,
            "error": f"Expected boolean, got {type(answer).__name__}",
        }

    if qtype == "scale":
        if isinstance(answer, (int, float)) and 1 <= answer <= 5:
            return {
                "id": qid,
                "type": qtype,
                "score": int(answer),
                "status": "scored",
                "raw_answer": answer,
            }
        return {
            "id": qid,
            "type": qtype,
            "score": None,
            "status": "invalid",
            "raw_answer": answer,
            "error": f"Expected integer 1-5, got {answer}",
        }

    if qtype == "text":
        has_content = isinstance(answer, str) and answer.strip()
        return {
            "id": qid,
            "type": qtype,
            "score": None,
            "status": "needs_review" if has_content else "unanswered",
            "raw_answer": answer if has_content else None,
        }

    return {
        "id": qid,
        "type": qtype,
        "score": None,
        "status": "unknown_type",
        "raw_answer": answer,
    }


def compute_criterion_score(criterion_scores: list[dict]) -> dict:
    """Compute average score for a criterion from its question scores."""
    scored = [s for s in criterion_scores if s["score"] is not None]
    needs_review = [s for s in criterion_scores if s["status"] == "needs_review"]

    if not scored:
        return {
            "score": None,
            "level": None,
            "scored_count": 0,
            "total_count": len(criterion_scores),
            "needs_review_count": len(needs_review),
        }

    avg = sum(s["score"] for s in scored) / len(scored)
    level = round(avg)

    level_names = {1: "Initial", 2: "Repeatable", 3: "Defined", 4: "Managed", 5: "Optimized"}

    return {
        "score": round(avg, 2),
        "level": level,
        "level_name": level_names.get(level, "Unknown"),
        "scored_count": len(scored),
        "total_count": len(criterion_scores),
        "needs_review_count": len(needs_review),
    }


def determine_tier(tier_scores: dict) -> dict:
    """Determine the organization's current maturity tier.

    An org is at Tier N if all criteria in Tiers 0..N score >= 3 (Defined).
    Tiers are evaluated in order; enrichment tiers are excluded from tier determination.
    """
    core_tier_order = ["tier_0", "tier_1", "tier_2", "tier_3", "tier_4"]
    achieved_tier = -1
    achieved_tier_name = "Below Foundation"

    tier_names = {
        "tier_0": "Tier 0: Foundation",
        "tier_1": "Tier 1: Basic",
        "tier_2": "Tier 2: Intermediate",
        "tier_3": "Tier 3: Advanced",
        "tier_4": "Tier 4: Expert",
    }

    for i, tier_id in enumerate(core_tier_order):
        if tier_id not in tier_scores:
            break

        tier_data = tier_scores[tier_id]
        all_defined = all(
            c["score"] is not None and c["score"] >= 3.0
            for c in tier_data["criteria"].values()
        )

        if all_defined:
            achieved_tier = i
            achieved_tier_name = tier_names[tier_id]
        else:
            break

    return {
        "tier_number": achieved_tier,
        "tier_name": achieved_tier_name,
        "description": (
            f"All criteria through {achieved_tier_name} meet or exceed Defined (3.0) level."
            if achieved_tier >= 0
            else "Not all Tier 0 (Foundation) criteria meet Defined (3.0) level."
        ),
    }


def run_scoring(rubric_path: Path, questionnaire_path: Path, response_path: Path) -> dict:
    """Run the full scoring pipeline. Returns structured results."""
    rubric = load_yaml(rubric_path)
    questionnaire = load_yaml(questionnaire_path)
    response = load_yaml(response_path)

    q_index = build_question_index(questionnaire)
    c_index = build_criterion_index(rubric)

    responses = response.get("responses", {})
    metadata = response.get("metadata", {})

    # Score each question
    question_scores = {}
    for qid, q_def in q_index.items():
        resp = responses.get(qid, {})
        answer = resp.get("answer") if isinstance(resp, dict) else resp
        question_scores[qid] = score_question(q_def, answer)
        question_scores[qid]["criterion"] = q_def["criterion"]
        question_scores[qid]["tier"] = q_def["tier"]

    # Group scores by criterion
    criterion_groups = {}
    for qid, qs in question_scores.items():
        crit = qs["criterion"]
        criterion_groups.setdefault(crit, []).append(qs)

    # Compute criterion scores
    criterion_results = {}
    for crit_id, scores in criterion_groups.items():
        crit_meta = c_index.get(crit_id, {})
        result = compute_criterion_score(scores)
        result["name"] = crit_meta.get("name", crit_id)
        result["weight"] = crit_meta.get("weight", 1.0)
        result["tier_id"] = crit_meta.get("tier_id", "unknown")
        result["tier_name"] = crit_meta.get("tier_name", "Unknown")
        result["questions"] = scores
        criterion_results[crit_id] = result

    # Group by tier
    tier_scores = {}
    for tier in rubric["tiers"]:
        tid = tier["id"]
        tier_criteria = {}
        scores_for_avg = []

        for crit in tier["criteria"]:
            cid = crit["id"]
            if cid in criterion_results:
                tier_criteria[cid] = criterion_results[cid]
                if criterion_results[cid]["score"] is not None:
                    scores_for_avg.append(
                        criterion_results[cid]["score"] * criterion_results[cid]["weight"]
                    )

        total_weight = sum(
            criterion_results[c["id"]]["weight"]
            for c in tier["criteria"]
            if c["id"] in criterion_results and criterion_results[c["id"]]["score"] is not None
        )

        tier_avg = round(sum(scores_for_avg) / total_weight, 2) if total_weight > 0 else None

        tier_scores[tid] = {
            "name": tier["name"],
            "score": tier_avg,
            "criteria": tier_criteria,
        }

    # Overall score (weighted average of all scored criteria)
    all_weighted = []
    total_weight = 0
    for crit_id, crit_data in criterion_results.items():
        if crit_data["score"] is not None:
            all_weighted.append(crit_data["score"] * crit_data["weight"])
            total_weight += crit_data["weight"]

    overall_score = round(sum(all_weighted) / total_weight, 2) if total_weight > 0 else None

    # Tier determination
    tier_determination = determine_tier(tier_scores)

    # Collect items needing review
    needs_review = [
        {
            "id": qs["id"],
            "criterion": qs["criterion"],
            "answer": qs["raw_answer"],
        }
        for qs in question_scores.values()
        if qs["status"] == "needs_review"
    ]

    # Collect issues (unanswered, invalid)
    issues = [
        {
            "id": qs["id"],
            "status": qs["status"],
            "error": qs.get("error"),
        }
        for qs in question_scores.values()
        if qs["status"] in ("unanswered", "invalid")
    ]

    # Auto-generate recommendations based on lowest scores
    recommendations = generate_recommendations(tier_scores, criterion_results)

    return {
        "metadata": metadata,
        "overall_score": overall_score,
        "tier_determination": tier_determination,
        "tier_scores": tier_scores,
        "needs_review": needs_review,
        "issues": issues,
        "recommendations": recommendations,
        "question_count": len(question_scores),
        "scored_count": sum(1 for q in question_scores.values() if q["status"] == "scored"),
        "needs_review_count": len(needs_review),
        "issue_count": len(issues),
    }


def generate_recommendations(tier_scores: dict, criterion_results: dict) -> list[dict]:
    """Generate improvement recommendations based on lowest-scoring areas."""
    recommendations = []

    # Find criteria scoring below Defined (3.0)
    below_defined = []
    for crit_id, crit_data in criterion_results.items():
        if crit_data["score"] is not None and crit_data["score"] < 3.0:
            below_defined.append(crit_data)

    # Sort by score (lowest first) then by tier order
    tier_order = {
        "tier_0": 0, "tier_1": 1, "tier_2": 2,
        "tier_3": 3, "tier_4": 4,
        "enrichment_people": 5, "enrichment_process": 6,
    }
    below_defined.sort(key=lambda c: (tier_order.get(c["tier_id"], 99), c["score"] or 0))

    for crit in below_defined:
        level_name = crit.get("level_name", "Unknown")
        target = "Defined" if crit["score"] < 3 else "Managed"
        recommendations.append({
            "criterion": crit["name"],
            "tier": crit["tier_name"],
            "current_score": crit["score"],
            "current_level": level_name,
            "recommendation": (
                f"'{crit['name']}' is at {level_name} ({crit['score']}/5.0). "
                f"Focus on reaching {target} level by establishing documented, "
                f"consistent processes in this area."
            ),
            "priority": "high" if crit["tier_id"] in ("tier_0", "tier_1") else "medium",
        })

    return recommendations


def print_results_rich(results: dict):
    """Print results using rich for terminal output."""
    try:
        from rich.console import Console
        from rich.table import Table
        from rich.panel import Panel
        from rich import box
    except ImportError:
        print_results_plain(results)
        return

    console = Console()

    # Header
    meta = results["metadata"]
    org = meta.get("organization", "Unknown")
    assessor = meta.get("assessor_name", "Unknown")
    date = meta.get("date", "Unknown")
    atype = meta.get("assessment_type", "Unknown")

    console.print()
    console.print(Panel(
        f"[bold]{org}[/bold]\n"
        f"Assessor: {assessor} | Date: {date} | Type: {atype}",
        title="[bold blue]DEBMM Assessment Results[/bold blue]",
        border_style="blue",
    ))

    # Executive Summary
    tier = results["tier_determination"]
    overall = results["overall_score"]
    console.print()
    console.print(Panel(
        f"[bold]Overall Score:[/bold] {overall}/5.0\n"
        f"[bold]Achieved Tier:[/bold] {tier['tier_name']}\n"
        f"[bold]Questions Scored:[/bold] {results['scored_count']}/{results['question_count']}\n"
        f"[bold]Needs Review:[/bold] {results['needs_review_count']} text answers\n"
        f"[bold]Issues:[/bold] {results['issue_count']} unanswered/invalid",
        title="[bold green]Executive Summary[/bold green]",
        border_style="green",
    ))

    # Tier breakdown
    for tid, tdata in results["tier_scores"].items():
        table = Table(
            title=f"{tdata['name']} - {tdata['score']}/5.0" if tdata["score"] else f"{tdata['name']} - N/A",
            box=box.ROUNDED,
            show_lines=True,
        )
        table.add_column("Criterion", style="cyan", min_width=35)
        table.add_column("Score", justify="center", min_width=8)
        table.add_column("Level", justify="center", min_width=12)
        table.add_column("Review", justify="center", min_width=8)

        for cid, cdata in tdata["criteria"].items():
            score_str = f"{cdata['score']}" if cdata["score"] is not None else "N/A"
            level_str = cdata.get("level_name", "N/A") if cdata["score"] is not None else "N/A"

            # Color based on score
            if cdata["score"] is not None:
                if cdata["score"] >= 4.0:
                    score_style = "bold green"
                elif cdata["score"] >= 3.0:
                    score_style = "bold yellow"
                elif cdata["score"] >= 2.0:
                    score_style = "bold dark_orange"
                else:
                    score_style = "bold red"
            else:
                score_style = "dim"

            review_str = f"{cdata['needs_review_count']} pending" if cdata["needs_review_count"] > 0 else "-"

            table.add_row(
                cdata["name"],
                f"[{score_style}]{score_str}[/{score_style}]",
                level_str,
                review_str,
            )

        console.print()
        console.print(table)

    # Recommendations
    if results["recommendations"]:
        console.print()
        rec_table = Table(
            title="Recommendations",
            box=box.ROUNDED,
            show_lines=True,
        )
        rec_table.add_column("Priority", justify="center", min_width=8)
        rec_table.add_column("Criterion", min_width=30)
        rec_table.add_column("Current", justify="center", min_width=10)
        rec_table.add_column("Recommendation", min_width=50)

        for rec in results["recommendations"]:
            priority_style = "bold red" if rec["priority"] == "high" else "bold yellow"
            rec_table.add_row(
                f"[{priority_style}]{rec['priority'].upper()}[/{priority_style}]",
                rec["criterion"],
                f"{rec['current_score']} ({rec['current_level']})",
                rec["recommendation"],
            )

        console.print(rec_table)

    console.print()


def print_results_plain(results: dict):
    """Fallback plain text output."""
    meta = results["metadata"]
    print(f"\n{'='*60}")
    print(f"DEBMM Assessment Results")
    print(f"{'='*60}")
    print(f"Organization: {meta.get('organization', 'Unknown')}")
    print(f"Assessor: {meta.get('assessor_name', 'Unknown')}")
    print(f"Date: {meta.get('date', 'Unknown')}")
    print(f"Type: {meta.get('assessment_type', 'Unknown')}")
    print(f"\nOverall Score: {results['overall_score']}/5.0")
    print(f"Achieved Tier: {results['tier_determination']['tier_name']}")
    print(f"Questions Scored: {results['scored_count']}/{results['question_count']}")
    print(f"Needs Review: {results['needs_review_count']} text answers")

    for tid, tdata in results["tier_scores"].items():
        print(f"\n--- {tdata['name']} ({tdata['score']}/5.0) ---")
        for cid, cdata in tdata["criteria"].items():
            score_str = f"{cdata['score']}" if cdata["score"] is not None else "N/A"
            level_str = cdata.get("level_name", "N/A") if cdata["score"] is not None else "N/A"
            print(f"  {cdata['name']}: {score_str} ({level_str})")

    if results["recommendations"]:
        print(f"\n--- Recommendations ---")
        for rec in results["recommendations"]:
            print(f"  [{rec['priority'].upper()}] {rec['recommendation']}")


def main():
    parser = argparse.ArgumentParser(
        description="Score a DEBMM assessment response.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("response", type=Path, nargs="?", help="Path to completed response YAML")
    parser.add_argument("--from-xlsx", type=Path, help="Score from a filled-out Excel spreadsheet")
    parser.add_argument("--rubric", type=Path, default=DEFAULT_RUBRIC, help="Path to rubric YAML")
    parser.add_argument(
        "--questionnaire", type=Path, default=DEFAULT_QUESTIONNAIRE,
        help="Path to questionnaire YAML",
    )
    parser.add_argument("--json", action="store_true", help="Output raw JSON results")
    parser.add_argument("--report", type=Path, help="Generate markdown report to file")
    parser.add_argument("--json-out", type=Path, help="Save JSON results to file")

    args = parser.parse_args()

    if not args.response and not args.from_xlsx:
        parser.error("Provide a response YAML file or use --from-xlsx with an Excel file.")

    if not args.rubric.exists():
        print(f"Error: Rubric file not found: {args.rubric}", file=sys.stderr)
        sys.exit(1)
    if not args.questionnaire.exists():
        print(f"Error: Questionnaire file not found: {args.questionnaire}", file=sys.stderr)
        sys.exit(1)

    if args.from_xlsx:
        if not args.from_xlsx.exists():
            print(f"Error: Excel file not found: {args.from_xlsx}", file=sys.stderr)
            sys.exit(1)
        response_data = load_xlsx_responses(args.from_xlsx, args.questionnaire)
        import tempfile
        with tempfile.NamedTemporaryFile(mode="w", suffix=".yaml", delete=False, encoding="utf-8") as f:
            yaml.dump(response_data, f, default_flow_style=False)
            temp_path = Path(f.name)
        try:
            results = run_scoring(args.rubric, args.questionnaire, temp_path)
        finally:
            temp_path.unlink()
    else:
        if not args.response.exists():
            print(f"Error: Response file not found: {args.response}", file=sys.stderr)
            sys.exit(1)
        results = run_scoring(args.rubric, args.questionnaire, args.response)

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2, default=str)
        print(f"JSON results saved to: {args.json_out}")

    if args.report:
        from report import generate_report
        report_md = generate_report(results)
        with open(args.report, "w", encoding="utf-8") as f:
            f.write(report_md)
        print(f"Report saved to: {args.report}")

    if args.json:
        print(json.dumps(results, indent=2, default=str))
    elif not args.report and not args.json_out:
        print_results_rich(results)


if __name__ == "__main__":
    main()
